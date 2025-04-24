import customtkinter as ctk
from openpyxl import Workbook, load_workbook
import os
from tkinter import ttk

archivo_guardar = "Ayuntamiento.xlsx"
fila_seleccionada = None

def guardar_excel():
    nombre = entry_nombre.get()
    apellido = entry_apellido.get()
    telefono = entry_telefono.get()
    if nombre == "" or telefono == "" or apellido == "":
        label.configure(text="campo vacio, necesita colocar info")
        return
    if not os.path.exists(archivo_guardar):
        wb = Workbook ()
        ws = wb.active
        ws.title = "Datos"
        ws.append(["Nombre", "Apellido", "Telefono"])
    else:
        wb = load_workbook(archivo_guardar)
        ws = wb.active
    
    #agregar un dato
    ws.append([nombre, apellido, telefono])
    wb.save(archivo_guardar)
    label.configure(text="dato guardado")
    entry_nombre.delete(0, ctk.END)
    entry_telefono.delete(0, ctk.END)
    entry_apellido.delete(0, ctk.END)

def actualizar_excel():
    global fila_seleccionada
    if fila_seleccionada is None:
        return

    nombre = entry_nombre.get()
    apellido = entry_apellido.get()
    telefono = entry_telefono.get()

    if nombre == "" or telefono == "" or apellido == "":
        label_estado.configure(text="Completá ambos campos.")
        return

    wb = load_workbook(archivo_guardar)
    ws = wb.active

    # Sobrescribimos la fila seleccionada
    ws.cell(row=fila_seleccionada, column=1, value=nombre)
    ws.cell(row=fila_seleccionada, column=2, value=apellido)
    ws.cell(row=fila_seleccionada, column=3, value=telefono)
    wb.save(archivo_guardar)

    label_estado.configure(text="✅ Registro actualizado")
    entry_nombre.delete(0, ctk.END)
    entry_apellido.delete(0, ctk.END)
    entry_telefono.delete(0, ctk.END)
    fila_seleccionada = None
    label_estado.configure(text="✏️ Editando registro...")
    boton_guardar.configure(text="Guardar", command=guardar_excel)


def mostrar_datos():
    for fila in tree.get_children():
        tree.delete(fila)
    if not os.path.exists(archivo_guardar):
        return
    
    wb = load_workbook(archivo_guardar)
    ws = wb.active

    for fila in ws.iter_rows(min_row=2, values_only=True):
        tree.insert("", "end", values=fila)
    
    
    tree.bind("<Double-1>", seleccionar_fila)


def cambiar_tabla():
    frame_registro.pack_forget()
    frame_tabla.pack(fill="both", expand=True)
    mostrar_datos()

def cambiar_registro():
    frame_tabla.pack_forget()
    frame_registro.pack(fill="both", expand=True)
    boton_guardar.configure(text="guardar datos", command=guardar_excel)
    label_estado.configure(text="")
    entry_nombre.delete(0, ctk.END)
    entry_apellido.delete(0, ctk.END)
    entry_telefono.delete(0, ctk.END)
    global fila_seleccionada
    fila_seleccionada = None

def seleccionar_fila(event):
    global fila_seleccionada
    item = tree.focus()
    valores = tree.item(item, "values")
    if not valores:
        return
    nombre, apellido, telefono = valores

    cambiar_registro()

    entry_nombre.delete(0, ctk.END)
    entry_nombre.insert(0, nombre)
    entry_apellido.delete(0, ctk.END)
    entry_apellido.insert(0, apellido)
    entry_telefono.delete(0, ctk.END)
    entry_telefono.insert(0, telefono)

    fila_seleccionada = tree.index(item) + 2
    boton_guardar.configure(text="actualizar datos", command = actualizar_excel)

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("green")

app = ctk.CTk()
app.title("Ayuntamiento")
app.geometry("600x400")

#frame de registro

frame_registro = ctk.CTkFrame(app)
frame_registro.pack(fill="both", expand=True)

label = ctk.CTkLabel(frame_registro, text="Nombre")
label.pack(pady=5)
entry_nombre = ctk.CTkEntry(frame_registro, placeholder_text="tu nombre")
entry_nombre.pack(pady=5)

label_apellido = ctk.CTkLabel(frame_registro, text="apellido")
label_apellido.pack(pady=5)
entry_apellido = ctk.CTkEntry(frame_registro, placeholder_text="ejemp: Martinez")
entry_apellido.pack(pady=5)


label_tel = ctk.CTkLabel(frame_registro, text="Telefono")
label_tel.pack(pady=5)
entry_telefono = ctk.CTkEntry(frame_registro, placeholder_text="ejemplo: 734218361623")
entry_telefono.pack(pady=5)

label_estado = ctk.CTkLabel(frame_registro, text="")
label_estado.pack(pady=5)

boton_guardar = ctk.CTkButton(frame_registro, text="guardar datos", command=guardar_excel)
boton_guardar.pack(pady=10)


boton_mostrar = ctk.CTkButton(frame_registro, text="mostrar datos", command=lambda: cambiar_tabla())
boton_mostrar.pack(pady=10)


frame_tabla = ctk.CTkFrame(app)

tree = ttk.Treeview(frame_tabla, columns=("Nombre", "Apellido", "Telefono"), show="headings")
tree.heading("Nombre", text="Nombre")
tree.heading("Apellido", text="Apellido")
tree.heading("Telefono", text="Telefono")
tree.pack(fill="both", expand=True)

button_volver = ctk.CTkButton(frame_tabla, text="volver a registro", command=lambda: cambiar_registro())
button_volver.pack(pady=10)


app.mainloop()
